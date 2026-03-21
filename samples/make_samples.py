"""
リアルな Excel 方眼紙ドキュメントを生成するスクリプト。

「本物の方眼紙」のルール:
  - 全列幅 2.5 文字幅、全行高 15pt で均一グリッド
  - セル結合はほぼしない（文字が隣の空セルにあふれるだけ）
  - 改行は「1行下のセルに書く」
  - 横並び項目は「文字数分だけ列をスキップして次のセルに書く」
  - 表も結合なし、罫線で囲んだ個別セルに値を入れる
  - 背景色や太字で見出しを区別するが、凝ったスタイルは使わない

使い方:
    python samples/make_samples.py
"""

from __future__ import annotations

import io
from pathlib import Path

# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _w(ws, row: int, col: int, value: str, bold: bool = False,
       size: float | None = None, italic: bool = False,
       bg: str | None = None, border: bool = False) -> None:
    """1セルに書き込む（結合なし）。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    c = ws.cell(row=row, column=col, value=value)
    font_kwargs: dict = {"name": "MS Gothic"}
    if bold:
        font_kwargs["bold"] = True
    if size:
        font_kwargs["size"] = size
    if italic:
        font_kwargs["italic"] = True
    c.font = Font(**font_kwargs)
    c.alignment = Alignment(wrap_text=False)

    if bg:
        c.fill = PatternFill(fill_type="solid", fgColor=bg)

    if border:
        side = Side(style="thin")
        c.border = Border(left=side, right=side, top=side, bottom=side)


def _border_row(ws, row: int, cols: list[int]) -> None:
    """罫線だけ引く（値なし）。"""
    for col in cols:
        _w(ws, row, col, "", border=True)


# ---------------------------------------------------------------------------
# xlsx: 議事録スタイル（会議のメモを方眼紙に書いた感じ）
# ---------------------------------------------------------------------------

def build_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # シート1: 議事録
    ws1 = wb.active
    ws1.title = "議事録"
    _setup_grid(ws1, cols=60, rows=80)
    _build_minutes(ws1)

    # シート2: TODO一覧（箇条書きスタイル）
    ws2 = wb.create_sheet("TODO一覧")
    _setup_grid(ws2, cols=60, rows=50)
    _build_todo(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup_grid(ws, cols: int = 60, rows: int = 80) -> None:
    from openpyxl.utils import get_column_letter
    for i in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 2.5
    for i in range(1, rows + 1):
        ws.row_dimensions[i].height = 15


def _build_minutes(ws) -> None:  # type: ignore[no-untyped-def]
    # ---- タイトル行（太字・大きめフォント、結合なし・単セル）----
    _w(ws,  1, 2, "システム開発　定例会議　議事録", bold=True, size=14)

    # ---- 会議基本情報（ラベルと値を列ずらしで横並び）----
    # 「日時：」は B3、値は「日時：」の文字幅(4文字 ≒ col6) から始める
    _w(ws,  3, 2, "日時：",   bold=True)
    _w(ws,  3, 6, "2025年3月15日（土）　14:00～16:30")   # 値は列6から

    _w(ws,  4, 2, "場所：",   bold=True)
    _w(ws,  4, 6, "第2会議室（本社ビル3F）")

    _w(ws,  5, 2, "参加者：", bold=True)
    _w(ws,  5, 6, "田中（PL）、鈴木（SE）、佐藤（SE）、山田（QA）、木村（顧客）")

    _w(ws,  6, 2, "欠席：",   bold=True)
    _w(ws,  6, 6, "中村（SE）※体調不良")

    _w(ws,  7, 2, "作成者：", bold=True)
    _w(ws,  7, 6, "鈴木")

    # ---- 空行（row8 は空のまま）----

    # ---- セクション見出し（太字、背景なし、単セル）----
    _w(ws,  9, 2, "１．前回議事録確認", bold=True)

    # 内容は B10 から。「改行」は次の行に書く
    _w(ws, 10, 2, "前回（3/1）の議事録を確認した。全員異議なし。")
    _w(ws, 11, 2, "アクション項目の進捗を以下の通り確認した。")

    # 箇条書き：先頭の「・」は col3、テキストは col4 から（少し右）
    _w(ws, 12, 3, "・")
    _w(ws, 12, 4, "ER図レビュー（担当：鈴木）　→　完了")

    _w(ws, 13, 3, "・")
    _w(ws, 13, 4, "テスト仕様書ドラフト（担当：山田）　→　作業中（3/22完成予定）")

    _w(ws, 14, 3, "・")
    _w(ws, 14, 4, "顧客向けデモ環境準備（担当：佐藤）　→　未着手　★遅延")

    # ---- 空行 ----

    _w(ws, 16, 2, "２．今回の議題", bold=True)

    # 議題1
    _w(ws, 17, 2, "（1）在庫照会画面の仕様変更について", bold=True)
    _w(ws, 18, 2, "顧客（木村氏）より以下の変更要望が提出された。")

    # ネストした箇条書き（さらに右側の列から）
    _w(ws, 19, 3, "①")
    _w(ws, 19, 4, "検索結果の表示件数を50件から100件に変更する")

    _w(ws, 20, 3, "②")
    _w(ws, 20, 4, "ソート順のデフォルトを「更新日降順」から「商品コード昇順」に変更する")

    _w(ws, 21, 3, "③")
    _w(ws, 21, 4, "CSV出力ボタンを画面上部に追加する（現在は下部のみ）")

    _w(ws, 22, 2, "【決定事項】上記①②③すべて対応することで合意。担当：佐藤　期限：3/29")
    _w(ws, 23, 2, "　　　　　　ただし③はUIデザイン変更を伴うため、デザインレビューを3/22に実施する。")

    # 議題2
    _w(ws, 25, 2, "（2）テスト計画について", bold=True)
    _w(ws, 26, 2, "山田より現在のテスト計画を報告。")
    _w(ws, 27, 2, "単体テスト：4/5～4/18　　結合テスト：4/21～5/2　　システムテスト：5/7～5/16")
    # ↑ 横に長い内容。「単体テスト：」を col2 に書いて残りをそのまま続けて書く

    _w(ws, 28, 2, "懸念事項：外部システム（物流API）との結合テスト環境が未確定。")
    _w(ws, 29, 2, "　　　　　4/7までに顧客側で環境を用意してもらう必要がある。")
    _w(ws, 30, 2, "【アクション】木村氏が社内で確認し、3/22までに回答する。")

    # 議題3
    _w(ws, 32, 2, "（3）本番リリース日程の確認", bold=True)
    _w(ws, 33, 2, "現状の計画：5/26（月）リリース")
    _w(ws, 34, 2, "テスト期間の遅れが発生した場合：6/2（月）に延期も検討。")
    _w(ws, 35, 2, "【決定】現時点は5/26を維持。4/25の定例で改めて判断する。")

    # ---- 空行 ----

    _w(ws, 37, 2, "３．次回アクションアイテム", bold=True)

    # 罫線付きの簡易テーブル（セル結合なし、罫線で囲む）
    # ヘッダー行
    headers = ["No.", "内容", "担当", "期限"]
    header_cols = [2, 4, 26, 31]     # 各列の開始列
    for header, col in zip(headers, header_cols):
        _w(ws, 38, col, header, bold=True, border=True)

    # データ行（各セルに罫線）
    table_data = [
        ("1", "UIデザインレビュー準備（CSV出力ボタン位置）", "佐藤", "3/22"),
        ("2", "物流API結合テスト環境の社内確認", "木村", "3/22"),
        ("3", "テスト仕様書ドラフト完成", "山田", "3/22"),
        ("4", "デモ環境構築", "佐藤", "3/29"),
        ("5", "在庫照会画面　仕様変更対応", "佐藤", "3/29"),
    ]
    for i, (no, content, tanto, limit) in enumerate(table_data):
        row = 39 + i
        _w(ws, row, 2,  no,      border=True)
        _w(ws, row, 4,  content, border=True)
        _w(ws, row, 26, tanto,   border=True)
        _w(ws, row, 31, limit,   border=True)

    # ---- 空行 ----

    _w(ws, 45, 2, "４．次回定例", bold=True)
    _w(ws, 46, 2, "日時：")
    _w(ws, 46, 6, "2025年3月22日（土）　14:00～")
    _w(ws, 47, 2, "場所：")
    _w(ws, 47, 6, "第2会議室（本社ビル3F）　※前回と同じ")
    _w(ws, 48, 2, "議題：")
    _w(ws, 48, 6, "UIデザインレビュー / テスト計画詳細 / アクション確認")

    _w(ws, 50, 2, "以上", italic=True)


def _build_todo(ws) -> None:  # type: ignore[no-untyped-def]
    _w(ws,  1, 2, "TODO一覧　（2025年3月時点）", bold=True, size=13)

    _w(ws,  3, 2, "★優先度：高", bold=True)
    # 各行：番号を col2、内容を col4、担当を col30、状態を col36 に書く（結合なし）
    for i, (content, tanto, status) in enumerate([
        ("デモ環境構築（AWS EC2セットアップ含む）",       "佐藤", "未着手"),
        ("物流API仕様書の入手・読み込み",                 "鈴木", "作業中"),
        ("在庫照会画面　仕様変更対応（UI含む）",           "佐藤", "未着手"),
    ]):
        row = 4 + i
        _w(ws, row, 2, f"□ {i+1}.", bold=True)
        _w(ws, row, 4, content)
        _w(ws, row, 30, tanto)
        _w(ws, row, 36, status, bold=(status == "未着手"))

    _w(ws,  8, 2, "★優先度：中", bold=True)
    for i, (content, tanto, status) in enumerate([
        ("テスト仕様書ドラフト作成",                     "山田", "作業中"),
        ("パフォーマンステスト計画書作成",               "鈴木", "未着手"),
        ("ユーザーマニュアル初稿作成",                   "佐藤", "未着手"),
        ("コードレビュー（在庫調整モジュール）",         "鈴木", "完了"),
    ]):
        row = 9 + i
        _w(ws, row, 2, f"□ {i+1}.", bold=True)
        _w(ws, row, 4, content)
        _w(ws, row, 30, tanto)
        _w(ws, row, 36, status)

    _w(ws, 14, 2, "★優先度：低", bold=True)
    for i, (content, tanto, status) in enumerate([
        ("README更新（開発環境セットアップ手順）",       "鈴木", "未着手"),
        ("古いテストデータの整理・削除",                 "山田", "未着手"),
    ]):
        row = 15 + i
        _w(ws, row, 2, f"□ {i+1}.", bold=True)
        _w(ws, row, 4, content)
        _w(ws, row, 30, tanto)
        _w(ws, row, 36, status)

    _w(ws, 18, 2, "【完了済み】", bold=True)
    for i, content in enumerate([
        "ER図レビュー（3/14完了）",
        "DB基本設計書　初稿レビュー（3/10完了）",
        "画面遷移図の確定（3/8完了）",
    ]):
        _w(ws, 19 + i, 3, f"✓ {content}", italic=True)

    _w(ws, 24, 2, "備考：")
    _w(ws, 24, 6, "「未着手」のうち優先度高の項目は3/22定例で進捗確認する")
    _w(ws, 25, 6, "「完了」でも問題が見つかった場合はステータスを「差戻し」に変更すること")


# ---------------------------------------------------------------------------
# xls: 仕様書スタイル（詳細設計書のメモを方眼紙に書いた感じ）
# ---------------------------------------------------------------------------

def build_xls() -> bytes:
    import xlwt

    wb = xlwt.Workbook(encoding="utf-8")
    _build_xls_spec(wb)
    _build_xls_check(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_plain(name: str = "MS Gothic", bold: bool = False,
               size: int = 11, italic: bool = False):  # type: ignore[no-untyped-def]
    import xlwt
    st = xlwt.XFStyle()
    f = xlwt.Font()
    f.name = name
    f.bold = bold
    f.height = size * 20
    f.italic = italic
    st.font = f
    return st


def _xls_bordered():  # type: ignore[no-untyped-def]
    import xlwt
    st = xlwt.XFStyle()
    f = xlwt.Font()
    f.name = "MS Gothic"
    st.font = f
    borders = xlwt.Borders()
    borders.left   = xlwt.Borders.THIN
    borders.right  = xlwt.Borders.THIN
    borders.top    = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    st.borders = borders
    return st


def _xw(ws, row: int, col: int, value: str, style=None) -> None:  # type: ignore[no-untyped-def]
    if style is None:
        import xlwt
        style = xlwt.XFStyle()
    ws.write(row, col, value, style)


def _build_xls_spec(wb) -> None:  # type: ignore[no-untyped-def]
    import xlwt
    ws = wb.add_sheet("詳細設計メモ")

    plain  = _xls_plain()
    bold   = _xls_plain(bold=True)
    bold14 = _xls_plain(bold=True, size=14)
    bold12 = _xls_plain(bold=True, size=12)
    italic = _xls_plain(italic=True)
    border = _xls_bordered()

    for col in range(60):
        ws.col(col).width = int(2.5 * 256)
    for row in range(80):
        ws.row(row).height_mismatch = True
        ws.row(row).height = 15 * 20

    # タイトル（単セル、大きいフォント）
    _xw(ws, 0, 1, "在庫管理システム　詳細設計メモ", bold14)

    # 基本情報（ラベルを col1、値を col6 に書く）
    for i, (label, value) in enumerate([
        ("機能名：",   "在庫照会"),
        ("担当SE：",   "鈴木 花子"),
        ("作成日：",   "2025-03-10"),
        ("更新日：",   "2025-03-15"),
    ]):
        _xw(ws, 2 + i, 1, label, bold)
        _xw(ws, 2 + i, 6, value, plain)

    # セクション1
    _xw(ws, 7, 1, "１．処理概要", bold12)
    _xw(ws, 8, 1, "ユーザーが検索条件を入力し、在庫テーブルを検索して結果を一覧表示する。", plain)
    _xw(ws, 9, 1, "検索条件が未入力の場合は全件取得する（上限1000件）。", plain)
    _xw(ws, 10, 1, "結果0件の場合は「データが見つかりませんでした」と表示する。", plain)

    # セクション2: 入力項目（箇条書き、改行ごとに1行）
    _xw(ws, 12, 1, "２．入力項目", bold12)
    for i, (item, note) in enumerate([
        ("商品コード（JANコード）",   "部分一致検索　例：4901234"),
        ("商品名",                   "部分一致検索"),
        ("カテゴリ",                 "プルダウン選択　全カテゴリ・食品・日用品・電子部品…"),
        ("倉庫",                     "プルダウン選択　全倉庫・東京・大阪・名古屋"),
        ("在庫数（下限）",           "数値入力　0以上の整数"),
        ("在庫数（上限）",           "数値入力　0以上の整数"),
    ]):
        row = 13 + i
        _xw(ws, row, 2, f"・{item}",           plain)   # 項目名は col2
        _xw(ws, row, 20, f"（{note}）",         italic)  # 補足は col20（文字幅分とばす）

    # セクション3: 出力項目（簡易テーブル）
    _xw(ws, 20, 1, "３．出力項目（一覧表示）", bold12)

    # ヘッダ（罫線あり、各セルに個別に書く）
    for col, header in [(1, "No."), (3, "項目名"), (15, "DB列名"), (25, "型"), (30, "備考")]:
        ws.write(21, col, header, border)
    for i, (no, name, col_name, dtype, note) in enumerate([
        ("1", "商品コード",   "item_code",    "VARCHAR(20)", ""),
        ("2", "商品名",       "item_name",    "VARCHAR(100)", ""),
        ("3", "カテゴリ名",   "category_name","VARCHAR(50)", "マスタJOIN"),
        ("4", "倉庫名",       "warehouse_name","VARCHAR(50)","マスタJOIN"),
        ("5", "在庫数",       "quantity",     "INTEGER",     ""),
        ("6", "最終更新日",   "updated_at",   "TIMESTAMP",   "YYYY-MM-DD HH:mm"),
    ]):
        row = 22 + i
        for col, val in [(1, no), (3, name), (15, col_name), (25, dtype), (30, note)]:
            ws.write(row, col, val, border)

    # セクション4: 処理フロー（ナンバリング、1行ずつ）
    _xw(ws, 29, 1, "４．処理フロー", bold12)
    steps = [
        "①　画面ロード時に倉庫・カテゴリのプルダウンリストをDBから取得する",
        "②　ユーザーが検索条件を入力して「検索」ボタンを押す",
        "③　入力値のバリデーションを行う（数値チェック・文字列長チェック等）",
        "　　エラーがある場合は画面上部にエラーメッセージを表示して処理を中断する",
        "④　バリデーション正常の場合、在庫テーブル（T_STOCK）を検索する",
        "　　結合：商品マスタ（M_ITEM）、倉庫マスタ（M_WAREHOUSE）、カテゴリマスタ（M_CATEGORY）",
        "⑤　結果を画面に表示する（最大1000件、ページング：50件/ページ）",
        "⑥　ユーザーが「CSV出力」ボタンを押した場合、全件（最大1000件）をCSV出力する",
    ]
    for i, step in enumerate(steps):
        _xw(ws, 30 + i, 1, step, plain)

    # セクション5: 備考・メモ（雑然と書き連ねる感じ）
    _xw(ws, 39, 1, "５．備考・TODO", bold12)
    _xw(ws, 40, 1, "・パフォーマンス：インデックス設計は鈴木が別途確認する（item_code, warehouse_id）", plain)
    _xw(ws, 41, 1, "・セキュリティ：SQLインジェクション対策はプリペアドステートメントで対応済み", plain)
    _xw(ws, 42, 1, "・TODO：カテゴリのプルダウンが多い場合の表示方法を田中PLに確認する　★3/20までに", plain)
    _xw(ws, 43, 1, "・TODO：在庫0件の商品を検索結果に含めるか仕様が未確定　→　顧客確認待ち", plain)
    _xw(ws, 44, 1, "・参考：既存システムの同等画面は「在庫参照機能仕様書 v2.3」を参照のこと", italic)


def _build_xls_check(wb) -> None:  # type: ignore[no-untyped-def]
    import xlwt
    ws = wb.add_sheet("チェックリスト")

    plain  = _xls_plain()
    bold   = _xls_plain(bold=True)
    bold13 = _xls_plain(bold=True, size=13)
    italic = _xls_plain(italic=True)
    border = _xls_bordered()

    for col in range(60):
        ws.col(col).width = int(2.5 * 256)
    for row in range(60):
        ws.row(row).height_mismatch = True
        ws.row(row).height = 15 * 20

    _xw(ws, 0, 1, "設計レビューチェックリスト", bold13)

    _xw(ws, 2, 1, "機能名：", bold)
    _xw(ws, 2, 6, "在庫照会", plain)
    _xw(ws, 3, 1, "レビュー日：", bold)
    _xw(ws, 3, 6, "2025-03-18", plain)
    _xw(ws, 4, 1, "レビュアー：", bold)
    _xw(ws, 4, 6, "田中（PL）、山田（QA）", plain)

    _xw(ws, 6, 1, "【機能要件チェック】", bold)

    # チェック表（No. / 確認項目 / 結果 / コメント）
    for col, hdr in [(1, "No."), (3, "確認項目"), (22, "結果"), (26, "コメント")]:
        ws.write(7, col, hdr, border)

    check_items = [
        ("1",  "入力項目がすべて定義されているか",                        "OK",   ""),
        ("2",  "出力項目がすべて定義されているか",                        "OK",   ""),
        ("3",  "バリデーションルールが明記されているか",                   "OK",   ""),
        ("4",  "エラー時の画面動作が定義されているか",                     "OK",   ""),
        ("5",  "パフォーマンス要件（3秒以内）を満たす設計か",              "要確認", "インデックス未確定"),
        ("6",  "セキュリティ対策（SQLインジェクション等）が考慮されているか", "OK", ""),
        ("7",  "ページング仕様が明記されているか",                         "OK",   ""),
        ("8",  "CSV出力の文字コードが指定されているか",                    "NG",   "UTF-8 or Shift-JIS未定"),
        ("9",  "0件時の動作が定義されているか",                           "OK",   ""),
        ("10", "在庫0件商品の扱いが定義されているか",                     "NG",   "顧客確認待ち★"),
    ]
    for i, (no, item, result, comment) in enumerate(check_items):
        row = 8 + i
        for col, val in [(1, no), (3, item), (22, result), (26, comment)]:
            ws.write(row, col, val, border)

    _xw(ws, 19, 1, "【非機能要件チェック】", bold)

    for col, hdr in [(1, "No."), (3, "確認項目"), (22, "結果"), (26, "コメント")]:
        ws.write(20, col, hdr, border)

    nfr_items = [
        ("1", "レスポンスタイム要件が記載されているか",   "OK",   "3秒以内"),
        ("2", "同時接続数が考慮されているか",             "要確認", "負荷テスト計画待ち"),
        ("3", "ログ設計が考慮されているか",               "OK",   ""),
    ]
    for i, (no, item, result, comment) in enumerate(nfr_items):
        row = 21 + i
        for col, val in [(1, no), (3, item), (22, result), (26, comment)]:
            ws.write(row, col, val, border)

    _xw(ws, 25, 1, "総合評価：", bold)
    _xw(ws, 25, 7, "条件付き承認（NGおよび要確認項目が解消されたら再レビュー）", plain)

    _xw(ws, 27, 1, "指摘事項まとめ：", bold)
    _xw(ws, 28, 2, "【NG-1】CSV出力の文字コードを明確にすること　→　担当：鈴木　期限：3/22", plain)
    _xw(ws, 29, 2, "【NG-2】在庫0件商品の扱いについて顧客確認を取ること　→　担当：田中PL　期限：3/22", plain)
    _xw(ws, 30, 2, "【要確認】パフォーマンス設計についてインデックス設計書を作成して再提出すること", plain)

    _xw(ws, 32, 1, "次回レビュー予定：", bold)
    _xw(ws, 32, 9, "2025-03-25（火）　14:00～", plain)

    _xw(ws, 34, 1, "※ このチェックリストは設計レビュー完了後に設計書と一緒に保管すること", italic)


# ---------------------------------------------------------------------------
# エントリーポイント
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    out_dir = Path(__file__).parent

    xlsx_path = out_dir / "議事録_システム開発定例.xlsx"
    xlsx_path.write_bytes(build_xlsx())
    print(f"生成完了: {xlsx_path}")

    xls_path = out_dir / "詳細設計メモ_在庫照会.xls"
    xls_path.write_bytes(build_xls())
    print(f"生成完了: {xls_path}")
