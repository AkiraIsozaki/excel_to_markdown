"""CLI 引数のパース・バリデーション、変換パイプラインの起動。

依存可能: 全モジュール (reader/, parser/, renderer/, models.py)
"""

from __future__ import annotations

import argparse
import dataclasses
import json
import sys
from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from excel_to_markdown import __version__
from excel_to_markdown.models import DocElement, RawCell, TextBlock
from excel_to_markdown.parser.cell_grid import CellGrid
from excel_to_markdown.parser.merge_resolver import resolve
from excel_to_markdown.parser.structure_detector import detect
from excel_to_markdown.parser.table_detector import find_tables
from excel_to_markdown.reader.xlsx_reader import read_sheet
from excel_to_markdown.renderer.markdown_renderer import render

# デフォルト値
DEFAULT_BASE_FONT_SIZE: float = 11.0
DEFAULT_COL_WIDTH: float = 8.0
DEFAULT_ROW_HEIGHT: float = 15.0


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """CLI 引数を解析する。バリデーションエラーは argparse が処理。

    第 1 引数が "serve" の場合は Web UI サーバーモード、それ以外は従来の変換モード。
    """
    argv_list = list(argv) if argv is not None else sys.argv[1:]

    if argv_list and argv_list[0] == "serve":
        return _parse_serve_args(argv_list[1:])

    return _parse_convert_args(argv_list)


def _parse_convert_args(argv: list[str]) -> argparse.Namespace:
    """変換モードの引数を解析する。"""
    parser = argparse.ArgumentParser(
        prog="python -m excel_to_markdown",
        description="Excel方眼紙 (.xlsx/.xls) を Markdown に変換する",
    )
    parser.add_argument(
        "input",
        help="変換する .xlsx/.xls ファイルのパス、またはディレクトリ（バッチ変換）",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        metavar="OUTPUT",
        help="出力 .md ファイルパス（省略時: 入力と同名 .md）。ディレクトリ入力時は無視",
    )
    parser.add_argument(
        "--sheet",
        "-s",
        default=None,
        metavar="SHEET",
        help="シート名または 0-based インデックス（省略時: 全シートを統合）",
    )
    parser.add_argument(
        "--base-font-size",
        type=float,
        default=DEFAULT_BASE_FONT_SIZE,
        metavar="SIZE",
        help=f"見出し判定の基準フォントサイズ（デフォルト: {DEFAULT_BASE_FONT_SIZE}）",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="TextBlock リストを JSON 形式で stderr に出力",
    )
    parser.add_argument(
        "--diagram",
        action="store_true",
        default=False,
        help="図形変換モード: Excel図形・コネクタを Mermaid flowchart に変換する",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    ns = parser.parse_args(argv)
    ns.subcommand = "convert"
    return ns


def _parse_serve_args(argv: list[str]) -> argparse.Namespace:
    """serve サブコマンドの引数を解析する。"""
    parser = argparse.ArgumentParser(
        prog="python -m excel_to_markdown serve",
        description="ブラウザ型 Web UI サーバーを起動する",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=8000,
        help="サーバーのポート番号（デフォルト: 8000）",
    )
    parser.add_argument(
        "--no-browser",
        action="store_true",
        default=False,
        help="ブラウザを自動で開かない",
    )
    ns = parser.parse_args(argv)
    ns.subcommand = "serve"
    return ns


def serve(args: argparse.Namespace) -> int:
    """Web UI サーバーを起動する。"""
    try:
        import webbrowser

        import uvicorn
    except ImportError:
        print(
            "エラー: Web UI の起動には uvicorn が必要です。"
            "`pip install excel-to-markdown[web]` でインストールしてください",
            file=sys.stderr,
        )
        return 1

    from excel_to_markdown.web.app import create_app

    port: int = args.port
    url = f"http://127.0.0.1:{port}"
    print(f"Web UI を起動中: {url}")

    if not args.no_browser:
        webbrowser.open(url)

    app = create_app()
    uvicorn.run(app, host="127.0.0.1", port=port)
    return 0


def run(args: argparse.Namespace) -> int:
    """変換パイプライン全体を実行し、exit code を返す。

    ディレクトリが指定された場合はバッチ変換モードで動作する。
    """
    try:
        input_path = Path(args.input).resolve()

        # --diagram: 図形変換モード
        if getattr(args, "diagram", False):
            return _run_diagram(input_path, args)

        # バッチ変換モード
        if input_path.is_dir():
            return _run_batch(input_path, args)

        _validate_input(input_path)
        output_path = _resolve_output_path(input_path, args.output)
        return _convert_file(input_path, output_path, args)

    except FileNotFoundError as e:
        print(f"エラー: ファイルが見つかりません: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1
    except PermissionError:
        print("エラー: 出力ファイルに書き込めません", file=sys.stderr)
        return 1
    except Exception as e:  # noqa: BLE001
        print(f"予期しないエラーが発生しました: {e}", file=sys.stderr)
        return 2


def _run_batch(dir_path: Path, args: argparse.Namespace) -> int:
    """ディレクトリ配下の全 xlsx/xls を変換する。エラー時は stderr に出力して継続。"""
    targets = sorted(
        list(dir_path.glob("**/*.xlsx")) + list(dir_path.glob("**/*.xls"))
    )
    if not targets:
        print(
            f"警告: ディレクトリ内に変換対象ファイルが見つかりません: {dir_path}",
            file=sys.stderr,
        )
        return 0

    exit_code = 0
    for file_path in targets:
        output_path = file_path.with_suffix(".md")
        code = _convert_file(file_path, output_path, args)
        if code != 0:
            exit_code = code
    return exit_code


def _convert_file(input_path: Path, output_path: Path, args: argparse.Namespace) -> int:
    """単一ファイルを変換して出力ファイルに書き出す。exit code を返す。"""
    try:
        sheet_markdowns = _process_workbook(input_path, args)

        if not sheet_markdowns:
            return 0

        if len(sheet_markdowns) == 1:
            final_md = sheet_markdowns[0][1]
        else:
            parts: list[str] = [
                f"# {sheet_name}\n\n{md}" for sheet_name, md in sheet_markdowns
            ]
            final_md = "\n\n---\n\n".join(parts)

        _write_output(output_path, final_md)
        return 0

    except FileNotFoundError as e:
        print(f"エラー: ファイルが見つかりません: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1
    except PermissionError:
        print(f"エラー: 出力ファイルに書き込めません: {output_path}", file=sys.stderr)
        return 1
    except Exception as e:  # noqa: BLE001
        print(f"予期しないエラーが発生しました ({input_path.name}): {e}", file=sys.stderr)
        return 2


def _process_workbook(
    input_path: Path, args: argparse.Namespace
) -> list[tuple[str, str]]:
    """ワークブックを処理して (sheet_name, markdown) のリストを返す。"""
    suffix = input_path.suffix.lower()
    if suffix == ".xls":
        return _process_xls(input_path, args)
    return _process_xlsx(input_path, args)


def _process_xlsx(
    input_path: Path, args: argparse.Namespace
) -> list[tuple[str, str]]:
    """xlsx ファイルを処理して (sheet_name, markdown) のリストを返す。

    drawingを持つシートは自動的にセル内容+Mermaidの統合出力を行う。
    """
    from excel_to_markdown.drawing.extractor import extract_sheet_drawing_map
    from excel_to_markdown.models import DiagramConnector, DiagramShape

    wb = _open_workbook(input_path)
    # シート名 → (shapes, connectors) マッピング（drawingなしのシートは含まれない）
    drawing_map: dict[str, tuple[list[DiagramShape], list[DiagramConnector]]] = {}
    try:
        drawing_map = extract_sheet_drawing_map(input_path)
    except Exception:  # noqa: BLE001
        pass  # drawing抽出失敗時は通常変換にフォールバック

    sheets = _select_sheets(wb, args.sheet)
    sheet_markdowns: list[tuple[str, str]] = []

    for ws in sheets:
        sheet_name: str = ws.title
        raw_cells = read_sheet(ws)
        has_cells = any(c.value for c in raw_cells)

        if sheet_name in drawing_map:
            shapes, connectors = drawing_map[sheet_name]
            # drawing行スパン（0-based drawing → 1-based cell に変換）
            drawing_top = min(s.top_row for s in shapes) + 1 if shapes else 1
            drawing_bottom = max(s.bottom_row for s in shapes) + 1 if shapes else 1

            md = _convert_sheet_combined(
                ws, raw_cells, shapes, connectors,
                drawing_top, drawing_bottom, args,
            )
            if md.strip():
                sheet_markdowns.append((sheet_name, md))
            continue

        if not has_cells:
            print(
                f'警告: シート "{sheet_name}" にコンテンツがありません。スキップします',
                file=sys.stderr,
            )
            continue

        grid = _build_grid(ws, raw_cells)
        md = _run_pipeline(raw_cells, grid, args)
        sheet_markdowns.append((sheet_name, md))

    return sheet_markdowns


def _process_xls(
    input_path: Path, args: argparse.Namespace
) -> list[tuple[str, str]]:
    """xls ファイルを処理して (sheet_name, markdown) のリストを返す。"""
    try:
        import xlrd
    except ImportError as e:
        raise ValueError(
            ".xls ファイルの変換には xlrd が必要です。`pip install xlrd` でインストールしてください"
        ) from e

    from excel_to_markdown.reader.xls_reader import read_sheet_xls

    try:
        book = xlrd.open_workbook(str(input_path), formatting_info=True)
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypted" in msg or "protect" in msg:
            raise ValueError("パスワード保護されたファイルは変換できません") from e
        raise

    sheet_markdowns: list[tuple[str, str]] = []

    if args.sheet is not None:
        sheet_names = book.sheet_names()
        if args.sheet.isdigit():
            idx = int(args.sheet)
            if idx >= len(sheet_names):
                raise ValueError(
                    f"シートが見つかりません: {args.sheet}（存在するシート: {sheet_names}）"
                )
            target_sheets = [book.sheet_by_index(idx)]
            target_names = [sheet_names[idx]]
        else:
            if args.sheet not in sheet_names:
                raise ValueError(
                    f"シートが見つかりません: {args.sheet}（存在するシート: {sheet_names}）"
                )
            target_sheets = [book.sheet_by_name(args.sheet)]
            target_names = [args.sheet]
    else:
        target_sheets = [book.sheet_by_index(i) for i in range(book.nsheets)]
        target_names = book.sheet_names()

    for sheet, sheet_name in zip(target_sheets, target_names):
        raw_cells = read_sheet_xls(sheet, book)
        if not any(c.value for c in raw_cells):
            print(
                f'警告: シート "{sheet_name}" にコンテンツがありません。スキップします',
                file=sys.stderr,
            )
            continue

        grid = CellGrid(cells=raw_cells)
        md = _run_pipeline(raw_cells, grid, args)
        sheet_markdowns.append((sheet_name, md))

    return sheet_markdowns


def _run_pipeline(
    raw_cells: list[RawCell], grid: CellGrid, args: argparse.Namespace
) -> str:
    """RawCell リストと CellGrid を受け取り Markdown 文字列を返す。"""
    blocks = resolve(raw_cells)

    if args.debug:
        _dump_blocks_debug(blocks)

    tables, remaining = find_tables(blocks, grid)
    doc_elements = detect(remaining, grid, args.base_font_size)

    all_elements: list[DocElement] = sorted(
        list(tables) + doc_elements,
        key=lambda e: e.source_row,
    )

    footnotes = [e.comment_text for e in all_elements if e.comment_text]
    return render(all_elements, footnotes)


def run_file(
    input_path: Path,
    base_font_size: float = DEFAULT_BASE_FONT_SIZE,
) -> str:
    """1つの Excel ファイルを Markdown 文字列に変換して返す。

    web/app.py からも呼び出される共通ヘルパー。
    全シートを統合した Markdown を返す。変換失敗時は例外を送出する。
    """
    _validate_input(input_path)
    dummy_args = argparse.Namespace(
        sheet=None,
        debug=False,
        base_font_size=base_font_size,
    )
    sheet_markdowns = _process_workbook(input_path, dummy_args)

    if not sheet_markdowns:
        return ""

    if len(sheet_markdowns) == 1:
        return sheet_markdowns[0][1]

    parts: list[str] = [f"# {name}\n\n{md}" for name, md in sheet_markdowns]
    return "\n\n---\n\n".join(parts)


def main() -> None:
    """CLI エントリーポイント。"""
    args = parse_args()
    if args.subcommand == "serve":
        sys.exit(serve(args))
    else:
        sys.exit(run(args))


# ---------------------------------------------------------------------------
# 内部ヘルパー
# ---------------------------------------------------------------------------


def _convert_sheet_combined(
    ws: Worksheet,
    raw_cells: list[RawCell],
    shapes: list,
    connectors: list,
    drawing_top_row: int,
    drawing_bottom_row: int,
    args: argparse.Namespace,
) -> str:
    """セル内容とDrawingを統合してMarkdown文字列を返す。

    drawing_top_row / drawing_bottom_row は 1-based（openpyxl のセル行番号基準）。

    出力構造:
        [drawing_top_row より前のセル → markdown]
        [Drawing → Mermaid block]
        [drawing_bottom_row より後のセル → markdown]
    """
    from excel_to_markdown.renderer.mermaid_renderer import render_mermaid_block

    parts: list[str] = []

    if any(c.value for c in raw_cells):
        grid = _build_grid(ws, raw_cells)
        blocks = resolve(raw_cells)

        if args.debug:
            _dump_blocks_debug(blocks)

        tables, remaining = find_tables(blocks, grid)
        doc_elements = detect(remaining, grid, args.base_font_size)
        all_elements: list[DocElement] = sorted(
            list(tables) + doc_elements, key=lambda e: e.source_row
        )

        # drawing 行範囲より前の要素
        before = [e for e in all_elements if e.source_row < drawing_top_row]
        # drawing 行範囲より後の要素
        after = [e for e in all_elements if e.source_row > drawing_bottom_row]

        if before:
            fn = [e.comment_text for e in before if e.comment_text]
            parts.append(render(before, fn))

    # Mermaid ブロック
    if shapes:
        parts.append(render_mermaid_block(shapes, connectors))

    if any(c.value for c in raw_cells):
        if after:  # type: ignore[possibly-undefined]
            fn2 = [e.comment_text for e in after if e.comment_text]
            parts.append(render(after, fn2))

    return "\n\n".join(p.strip() for p in parts if p.strip()) + "\n"


def _run_diagram(input_path: Path, args: argparse.Namespace) -> int:
    """--diagram モード: Excel図形をMermaid形式に変換して出力する。"""
    from excel_to_markdown.drawing.extractor import extract_diagrams
    from excel_to_markdown.renderer.mermaid_renderer import render_mermaid_block

    _validate_input(input_path)

    results = extract_diagrams(input_path)
    if not results:
        print(
            f"警告: {input_path.name} に図形（Drawing）が見つかりませんでした",
            file=sys.stderr,
        )
        return 0

    blocks: list[str] = []
    for idx, (shapes, connectors) in enumerate(results, 1):
        if not shapes:
            continue
        block = render_mermaid_block(shapes, connectors)
        if len(results) > 1:
            blocks.append(f"## Drawing {idx}\n\n{block}")
        else:
            blocks.append(block)

    if not blocks:
        print(
            f"警告: {input_path.name} の Drawing に図形がありませんでした",
            file=sys.stderr,
        )
        return 0

    output = "\n".join(blocks)

    output_path_arg: str | None = getattr(args, "output", None)
    if output_path_arg:
        out = Path(output_path_arg).resolve()
        _write_output(out, output)
        print(f"Mermaid を出力しました: {out}")
    else:
        print(output)

    return 0


def _validate_input(path: Path) -> None:
    """入力ファイルのバリデーション。"""
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise ValueError(
            f"対応していないファイル形式です: {path.suffix}（.xlsx/.xls のみ対応）"
        )


def _resolve_output_path(input_path: Path, output_arg: str | None) -> Path:
    """出力ファイルパスを決定する。"""
    if output_arg is not None:
        return Path(output_arg).resolve()
    return input_path.with_suffix(".md")


def _open_workbook(path: Path) -> openpyxl.Workbook:
    """ワークブックを開く。パスワード保護されている場合は ValueError を送出。"""
    try:
        return openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypted" in msg or "protect" in msg:
            raise ValueError("パスワード保護されたファイルは変換できません") from e
        raise


def _select_sheets(
    wb: openpyxl.Workbook, sheet_arg: str | None
) -> list[Worksheet]:
    """対象シートのリストを返す。"""
    if sheet_arg is None:
        return [wb[name] for name in wb.sheetnames]

    # 数値インデックス (0-based)
    if sheet_arg.isdigit():
        idx = int(sheet_arg)
        if idx >= len(wb.sheetnames):
            raise ValueError(
                f"シートが見つかりません: {sheet_arg}"
                f"（存在するシート: {wb.sheetnames}）"
            )
        return [wb[wb.sheetnames[idx]]]

    # シート名で検索
    if sheet_arg not in wb.sheetnames:
        raise ValueError(
            f"シートが見つかりません: {sheet_arg}"
            f"（存在するシート: {wb.sheetnames}）"
        )
    return [wb[sheet_arg]]


def _build_grid(ws: Worksheet, raw_cells: list[RawCell]) -> CellGrid:
    """ワークシートから CellGrid を構築する。"""
    col_widths: dict[int, float] = {
        column_index_from_string(col): ws.column_dimensions[col].width or DEFAULT_COL_WIDTH
        for col in ws.column_dimensions
    }
    row_heights: dict[int, float] = {
        r: ws.row_dimensions[r].height or DEFAULT_ROW_HEIGHT
        for r in ws.row_dimensions
    }
    return CellGrid(cells=raw_cells, col_widths=col_widths, row_heights=row_heights)


def _dump_blocks_debug(blocks: list[TextBlock]) -> None:
    """TextBlock リストを JSON 形式で stderr に出力する。"""
    data = [dataclasses.asdict(b) for b in blocks]
    print(json.dumps(data, ensure_ascii=False, indent=2), file=sys.stderr)


def _write_output(path: Path, content: str) -> None:
    """Markdown を UTF-8 でアトミックに書き出す。

    tmpファイルに書き込んでから rename することで、
    書き込み失敗時に部分的な出力ファイルが残らないことを保証する。
    """
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    try:
        tmp_path.write_text(content, encoding="utf-8")
        tmp_path.replace(path)
    except OSError as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise PermissionError(f"出力ファイルに書き込めません: {path}") from e
