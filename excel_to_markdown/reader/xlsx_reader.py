"""openpyxl を使って .xlsx シートから list[RawCell] を抽出する。

依存可能: models.py, openpyxl
依存禁止: parser/, renderer/, cli.py
"""

from __future__ import annotations

from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from excel_to_markdown.models import RawCell

# openpyxl のリッチテキスト型（インポートに失敗しても動作するよう try-except）
try:
    from openpyxl.cell.rich_text import CellRichText as _CellRichText
except ImportError:  # pragma: no cover
    _CellRichText = None  # type: ignore[assignment,misc]


def read_sheet(ws: Worksheet, print_area: str | None = None) -> list[RawCell]:
    """ワークシートから RawCell リストを返す。

    - 印刷領域が設定されている場合はその範囲内のみ処理
    - 非表示の行・列は除外
    - 結合セルの非起点セルは is_merge_origin=False, value=None でスキップ対象
    """
    # 呼び出し元から文字列で印刷領域が渡された場合はそれを優先し、なければ ws から取得
    if print_area is not None:
        area = _parse_area_str(print_area)
    else:
        area = get_print_area(ws)  # (min_row, min_col, max_row, max_col) | None

    # 非表示行・列セットを事前収集
    hidden_rows: set[int] = {
        r for r, rd in ws.row_dimensions.items() if rd.hidden
    }
    hidden_cols: set[int] = {
        column_index_from_string(cd)  # 列文字 → 列番号 (1-based)
        for cd in ws.column_dimensions
        if ws.column_dimensions[cd].hidden
    }

    # 結合セル情報を事前収集: 結合起点 → (row_span, col_span)
    merge_origins: dict[tuple[int, int], tuple[int, int]] = {}
    merge_non_origins: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        origin = (mr.min_row, mr.min_col)
        row_span = mr.max_row - mr.min_row + 1
        col_span = mr.max_col - mr.min_col + 1
        merge_origins[origin] = (row_span, col_span)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != origin:
                    merge_non_origins.add((r, c))

    raw_cells: list[RawCell] = []

    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column

            # 印刷領域フィルタ
            if area is not None:
                min_r, min_c, max_r, max_c = area
                if not (min_r <= r <= max_r and min_c <= c <= max_c):
                    continue

            # 非表示行・列を除外
            if r in hidden_rows or c in hidden_cols:
                continue

            pos = (r, c)
            is_origin = pos in merge_origins
            is_non_origin = pos in merge_non_origins

            if is_non_origin:
                # 結合の非起点セルは value=None, is_merge_origin=False で記録
                raw_cells.append(
                    RawCell(
                        row=r,
                        col=c,
                        value=None,
                        font_bold=False,
                        font_italic=False,
                        font_strikethrough=False,
                        font_underline=False,
                        font_size=None,
                        font_color=None,
                        bg_color=None,
                        is_merge_origin=False,
                        merge_row_span=1,
                        merge_col_span=1,
                        has_comment=False,
                        comment_text=None,
                    )
                )
                continue

            # セル値の文字列化
            value = _cell_value_to_str(cell)

            font_bold, font_italic, font_strike, font_underline, font_size, font_color = (
                extract_font_props(cell)
            )
            bg_color = extract_bg_color(cell)

            row_span, col_span = merge_origins.get(pos, (1, 1))
            has_comment = cell.comment is not None
            comment_text = cell.comment.text if has_comment else None
            hyperlink: str | None = None
            if cell.hyperlink is not None:
                hyperlink = cell.hyperlink.target or None

            raw_cells.append(
                RawCell(
                    row=r,
                    col=c,
                    value=value,
                    font_bold=font_bold,
                    font_italic=font_italic,
                    font_strikethrough=font_strike,
                    font_underline=font_underline,
                    font_size=font_size,
                    font_color=font_color,
                    bg_color=bg_color,
                    is_merge_origin=is_origin,
                    merge_row_span=row_span,
                    merge_col_span=col_span,
                    has_comment=has_comment,
                    comment_text=comment_text,
                    hyperlink=hyperlink,
                )
            )

    return raw_cells


def _cell_value_to_str(cell: Cell) -> str | None:
    """セルの値を文字列に変換する。None / 空文字はNoneを返す。"""
    v = cell.value
    if v is None:
        return None
    # openpyxl のリッチテキスト型
    if _CellRichText is not None and isinstance(v, _CellRichText):
        text = str(v)
        return text if text.strip() else None
    text = str(v)
    return text if text.strip() else None


def extract_font_props(
    cell: Cell,
) -> tuple[bool, bool, bool, bool, float | None, str | None]:
    """セルのフォントプロパティ (bold, italic, strike, underline, size, color) を返す。

    テーマ色は解決せずNoneを返す。
    戻り値: (bold, italic, strikethrough, underline, size, color)
    """
    font = cell.font
    if font is None:
        return False, False, False, False, None, None

    bold: bool = bool(font.bold)
    italic: bool = bool(font.italic)
    strike: bool = bool(font.strike)
    underline: bool = bool(font.underline) and font.underline != "none"

    size: float | None = None
    if font.size is not None:
        try:
            size = float(font.size)
        except (TypeError, ValueError):
            size = None

    color: str | None = None
    if font.color is not None and font.color.type == "rgb":
        raw = font.color.rgb
        if raw and raw != "00000000":
            color = str(raw)

    return bold, italic, strike, underline, size, color


def extract_bg_color(cell: Cell) -> str | None:
    """セルの背景色を ARGB hex で返す。塗りつぶしなし・グラデーションはNoneを返す。"""
    fill = cell.fill
    if fill is None:
        return None
    if fill.fill_type in (None, "none", "gradient"):
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb":
        raw = str(fg.rgb)
        # 透明色 (00000000) はNone扱い
        return raw if raw != "00000000" else None
    return None


def get_print_area(ws: Worksheet) -> tuple[int, int, int, int] | None:
    """印刷領域を (min_row, min_col, max_row, max_col) で返す。未設定はNone。"""
    pa = ws.print_area
    if not pa:
        return None

    # print_area は "A1:D10" のような文字列、または複数範囲のリスト
    area_str: str
    if isinstance(pa, list):
        area_str = pa[0]
    else:
        area_str = str(pa)

    return _parse_area_str(area_str)


def _parse_area_str(area_str: str) -> tuple[int, int, int, int] | None:
    """セル範囲文字列 ("A1:D10" 等) を (min_row, min_col, max_row, max_col) に変換する。

    "SheetName!A1:D10" 形式にも対応する。解析失敗時はNoneを返す。
    """
    # "SheetName!A1:D10" 形式の場合、シート名部分を除去
    if "!" in area_str:
        area_str = area_str.split("!")[-1]

    try:
        from openpyxl.utils import range_boundaries

        min_c, min_r, max_c, max_r = range_boundaries(area_str)
        return min_r, min_c, max_r, max_c
    except Exception:
        return None
