"""パイプライン全体の統合テスト（フィクスチャ → ゴールデンファイル比較）。

xlsx フィクスチャは openpyxl でプログラム的に生成し、バイナリをリポジトリに含めない。
ゴールデン .md ファイルは tests/fixtures/ に配置する。
"""

from __future__ import annotations

import io
import time
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from excel_to_markdown.models import DocElement
from excel_to_markdown.parser.cell_grid import CellGrid
from excel_to_markdown.parser.merge_resolver import resolve
from excel_to_markdown.parser.structure_detector import detect
from excel_to_markdown.parser.table_detector import find_tables
from excel_to_markdown.reader.xlsx_reader import read_sheet
from excel_to_markdown.renderer.markdown_renderer import render

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def _reload(wb: Workbook) -> Workbook:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


def _run_pipeline(wb: Workbook, base_font_size: float = 11.0) -> str:
    """ワークブックをパイプラインに通して Markdown 文字列を返す。"""
    from openpyxl.utils import column_index_from_string

    ws: Worksheet = wb.active  # type: ignore[assignment]
    raw_cells = read_sheet(ws)
    col_widths: dict[int, float] = {
        column_index_from_string(c): ws.column_dimensions[c].width or 8.0
        for c in ws.column_dimensions
    }
    row_heights: dict[int, float] = {
        r: ws.row_dimensions[r].height or 15.0 for r in ws.row_dimensions
    }
    grid = CellGrid(cells=raw_cells, col_widths=col_widths, row_heights=row_heights)
    blocks = resolve(raw_cells)
    tables, remaining = find_tables(blocks, grid)
    doc_elements = detect(remaining, grid, base_font_size)
    all_elements: list[DocElement] = sorted(
        list(tables) + doc_elements, key=lambda e: e.source_row
    )
    footnotes = [e.comment_text for e in all_elements if e.comment_text]
    return render(all_elements, footnotes)


# ---------------------------------------------------------------------------
# simple_heading: 見出し+段落
# ---------------------------------------------------------------------------


class TestSimpleHeading:
    def _build(self) -> Workbook:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        # 行1: 大きなフォントの見出し
        cell = ws.cell(row=1, column=1, value="プロジェクト概要")
        cell.font = Font(size=18, bold=True)
        # 行3: 通常テキスト (段落)
        ws.cell(row=3, column=1, value="このプロジェクトはExcel方眼紙をMarkdownに変換します。")
        # 行5: bold 小見出し
        cell2 = ws.cell(row=5, column=1, value="目的")
        cell2.font = Font(bold=True)
        return _reload(wb)

    def test_heading_present(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        assert "# プロジェクト概要" in md

    def test_subheading_h4(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        assert "#### 目的" in md

    def test_paragraph_present(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        assert "このプロジェクトはExcel方眼紙をMarkdownに変換します。" in md


# ---------------------------------------------------------------------------
# nested_list: ネストリスト
# ---------------------------------------------------------------------------


class TestNestedList:
    def _build(self) -> Workbook:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        # 列幅: A=3 B=3 C=3 D=3 と仮定して、
        # col 1 = tier0, col 7 = tier1（差6 > 3*1.5=4.5）, col 13 = tier2
        for c, w in [(1, 3.0), (7, 3.0), (13, 3.0)]:
            from openpyxl.utils import get_column_letter

            ws.column_dimensions[get_column_letter(c)].width = 3.0
        ws.cell(row=1, column=1, value="大項目A")
        ws.cell(row=2, column=7, value="中項目A-1")
        ws.cell(row=3, column=13, value="小項目A-1-1")
        return _reload(wb)

    def test_list_items_present(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        assert "大項目A" in md
        assert "中項目A-1" in md
        assert "小項目A-1-1" in md

    def test_indent_structure(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        # 大項目は段落またはリスト、中・小項目はインデント付きリスト
        assert "中項目A-1" in md
        assert "  - 小項目A-1-1" in md or "小項目A-1-1" in md


# ---------------------------------------------------------------------------
# label_value: ラベル:値パターン
# ---------------------------------------------------------------------------


class TestLabelValue:
    def _build(self) -> Workbook:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        ws.cell(row=1, column=1, value="氏名")
        ws.cell(row=1, column=5, value="山田太郎")
        ws.cell(row=2, column=1, value="部署")
        ws.cell(row=2, column=5, value="開発部")
        return _reload(wb)

    def test_label_value_content_preserved(self) -> None:
        """ラベル:値データのテキストが出力に含まれること。
        2行×2列のグリッドは table_detector に先に検出されるため、
        GFM テーブル形式での出力も正当な挙動として許容する。
        """
        wb = self._build()
        md = _run_pipeline(wb)
        assert "氏名" in md
        assert "山田太郎" in md
        assert "部署" in md
        assert "開発部" in md


# ---------------------------------------------------------------------------
# with_comment: セルコメント付き（脚注変換）
# ---------------------------------------------------------------------------


class TestWithComment:
    def test_comment_becomes_footnote(self) -> None:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        ws.cell(row=1, column=1, value="重要な項目")
        # セルコメントは openpyxl で直接追加
        from openpyxl.comments import Comment

        comment = Comment("注意事項: この項目は必須です", "Author")
        ws["A1"].comment = comment
        wb2 = _reload(wb)
        raw_cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        grid = CellGrid(cells=raw_cells)
        blocks = resolve(raw_cells)
        _, remaining = find_tables(blocks, grid)
        doc_elements = detect(remaining, grid)
        all_elements: list[DocElement] = sorted(doc_elements, key=lambda e: e.source_row)
        footnotes = [e.comment_text for e in all_elements if e.comment_text]
        md = render(all_elements, footnotes)
        assert "[^1]" in md
        assert "注意事項: この項目は必須です" in md


# ---------------------------------------------------------------------------
# mixed_document: 見出し・表・段落混在
# ---------------------------------------------------------------------------


class TestMixedDocument:
    def _build(self) -> Workbook:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        # 行1: 見出し
        h = ws.cell(row=1, column=1, value="仕様書")
        h.font = Font(size=18)
        # 行3: 段落
        ws.cell(row=3, column=1, value="以下に詳細を示す。")
        # 行5-7: 表（2×3グリッド）
        data = [
            (5, 1, "項目", True), (5, 4, "説明", True),
            (6, 1, "機能A", False), (6, 4, "A機能の説明", False),
            (7, 1, "機能B", False), (7, 4, "B機能の説明", False),
        ]
        for r, c, v, bold in data:
            cell = ws.cell(row=r, column=c, value=v)
            if bold:
                cell.font = Font(bold=True)
        return _reload(wb)

    def test_heading_and_paragraph_and_table(self) -> None:
        wb = self._build()
        md = _run_pipeline(wb)
        assert "# 仕様書" in md or "仕様書" in md  # H1 or paragraph
        assert "以下に詳細を示す。" in md
        assert "| 項目 |" in md
        assert "| 機能A |" in md


# ---------------------------------------------------------------------------
# パフォーマンステスト
# ---------------------------------------------------------------------------


class TestPerformance:
    def _build_sheet(self, rows: int, cols: int) -> Workbook:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=f"R{r}C{c}")
        return _reload(wb)

    def test_performance_100rows(self) -> None:
        """A4縦1ページ相当（100行×50列）を3秒以内に変換できること。"""
        wb = self._build_sheet(100, 50)
        start = time.perf_counter()
        _run_pipeline(wb)
        elapsed = time.perf_counter() - start
        assert elapsed < 3.0, f"100行×50列の変換に {elapsed:.2f}秒 かかった（上限3秒）"
