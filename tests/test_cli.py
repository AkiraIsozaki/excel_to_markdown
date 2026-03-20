"""cli.py のユニットテスト。"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from excel_to_markdown.cli import (
    _resolve_output_path,
    _select_sheets,
    _validate_input,
    _write_output,
    parse_args,
    run,
)


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------


def _save_to_file(wb: Workbook, path: Path) -> None:
    """Workbook をファイルに保存する。"""
    buf = io.BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())


def _make_simple_xlsx(path: Path, sheet_name: str = "Sheet", value: str = "テスト") -> None:
    """1セルだけ持つシンプルな xlsx を生成して保存する。"""
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = sheet_name
    ws.cell(row=1, column=1, value=value)
    _save_to_file(wb, path)


# ---------------------------------------------------------------------------
# parse_args
# ---------------------------------------------------------------------------


class TestParseArgs:
    def test_input_only(self) -> None:
        args = parse_args(["input.xlsx"])
        assert args.input == "input.xlsx"
        assert args.output is None
        assert args.sheet is None
        assert args.base_font_size == 11.0
        assert args.debug is False

    def test_output_short(self) -> None:
        args = parse_args(["input.xlsx", "-o", "out.md"])
        assert args.output == "out.md"

    def test_output_long(self) -> None:
        args = parse_args(["input.xlsx", "--output", "out.md"])
        assert args.output == "out.md"

    def test_sheet_short(self) -> None:
        args = parse_args(["input.xlsx", "-s", "Sheet1"])
        assert args.sheet == "Sheet1"

    def test_sheet_long(self) -> None:
        args = parse_args(["input.xlsx", "--sheet", "0"])
        assert args.sheet == "0"

    def test_base_font_size(self) -> None:
        args = parse_args(["input.xlsx", "--base-font-size", "10.5"])
        assert args.base_font_size == 10.5

    def test_debug_flag(self) -> None:
        args = parse_args(["input.xlsx", "--debug"])
        assert args.debug is True

    def test_no_args_exits(self) -> None:
        with pytest.raises(SystemExit):
            parse_args([])


# ---------------------------------------------------------------------------
# _validate_input
# ---------------------------------------------------------------------------


class TestValidateInput:
    def test_valid_xlsx(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        f.write_bytes(b"dummy")
        _validate_input(f)  # 例外が出なければOK

    def test_file_not_found(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            _validate_input(tmp_path / "nonexistent.xlsx")

    def test_unsupported_extension(self, tmp_path: Path) -> None:
        f = tmp_path / "test.csv"
        f.write_bytes(b"dummy")
        with pytest.raises(ValueError, match="対応していないファイル形式"):
            _validate_input(f)

    def test_xls_is_valid(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xls"
        f.write_bytes(b"dummy")
        _validate_input(f)  # 例外が出なければOK


# ---------------------------------------------------------------------------
# _resolve_output_path
# ---------------------------------------------------------------------------


class TestResolveOutputPath:
    def test_output_not_specified_uses_same_name(self, tmp_path: Path) -> None:
        input_path = tmp_path / "report.xlsx"
        result = _resolve_output_path(input_path, None)
        assert result == input_path.with_suffix(".md")

    def test_output_specified(self, tmp_path: Path) -> None:
        input_path = tmp_path / "report.xlsx"
        out = str(tmp_path / "custom.md")
        result = _resolve_output_path(input_path, out)
        assert result == Path(out).resolve()


# ---------------------------------------------------------------------------
# _select_sheets
# ---------------------------------------------------------------------------


class TestSelectSheets:
    def _make_wb(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"  # type: ignore[union-attr]
        wb.create_sheet("Sheet2")
        return wb

    def test_no_sheet_arg_returns_all(self) -> None:
        wb = self._make_wb()
        sheets = _select_sheets(wb, None)
        assert len(sheets) == 2
        assert sheets[0].title == "Sheet1"
        assert sheets[1].title == "Sheet2"

    def test_select_by_name(self) -> None:
        wb = self._make_wb()
        sheets = _select_sheets(wb, "Sheet2")
        assert len(sheets) == 1
        assert sheets[0].title == "Sheet2"

    def test_select_by_index_0(self) -> None:
        wb = self._make_wb()
        sheets = _select_sheets(wb, "0")
        assert sheets[0].title == "Sheet1"

    def test_select_by_index_1(self) -> None:
        wb = self._make_wb()
        sheets = _select_sheets(wb, "1")
        assert sheets[0].title == "Sheet2"

    def test_index_out_of_range(self) -> None:
        wb = self._make_wb()
        with pytest.raises(ValueError, match="シートが見つかりません"):
            _select_sheets(wb, "99")

    def test_sheet_name_not_found(self) -> None:
        wb = self._make_wb()
        with pytest.raises(ValueError, match="シートが見つかりません"):
            _select_sheets(wb, "NoSuchSheet")


# ---------------------------------------------------------------------------
# _write_output
# ---------------------------------------------------------------------------


class TestWriteOutput:
    def test_writes_content(self, tmp_path: Path) -> None:
        out = tmp_path / "out.md"
        _write_output(out, "# Hello\n")
        assert out.read_text(encoding="utf-8") == "# Hello\n"

    def test_permission_error_on_readonly_dir(self, tmp_path: Path) -> None:
        readonly_dir = tmp_path / "readonly"
        readonly_dir.mkdir()
        readonly_dir.chmod(0o555)
        out = readonly_dir / "out.md"
        try:
            with pytest.raises(PermissionError):
                _write_output(out, "content")
        finally:
            readonly_dir.chmod(0o755)


# ---------------------------------------------------------------------------
# run() 正常系
# ---------------------------------------------------------------------------


class TestRunSuccess:
    def test_single_sheet_creates_md(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "doc.xlsx"
        _make_simple_xlsx(xlsx, value="見出し")
        args = parse_args([str(xlsx)])
        code = run(args)
        assert code == 0
        md_path = xlsx.with_suffix(".md")
        assert md_path.exists()
        assert "見出し" in md_path.read_text(encoding="utf-8")

    def test_output_path_specified(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "doc.xlsx"
        out = tmp_path / "result.md"
        _make_simple_xlsx(xlsx, value="テスト内容")
        args = parse_args([str(xlsx), "-o", str(out)])
        code = run(args)
        assert code == 0
        assert out.exists()

    def test_multiple_sheets_merged(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws1: Worksheet = wb.active  # type: ignore[assignment]
        ws1.title = "概要"
        ws1.cell(row=1, column=1, value="概要シートの内容")
        ws2: Worksheet = wb.create_sheet("詳細")
        ws2.cell(row=1, column=1, value="詳細シートの内容")
        xlsx = tmp_path / "multi.xlsx"
        _save_to_file(wb, xlsx)

        args = parse_args([str(xlsx)])
        code = run(args)
        assert code == 0
        md = xlsx.with_suffix(".md").read_text(encoding="utf-8")
        assert "概要シートの内容" in md
        assert "詳細シートの内容" in md
        assert "# 概要" in md
        assert "# 詳細" in md

    def test_sheet_name_selection(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws1: Worksheet = wb.active  # type: ignore[assignment]
        ws1.title = "SheetA"
        ws1.cell(row=1, column=1, value="A内容")
        ws2: Worksheet = wb.create_sheet("SheetB")
        ws2.cell(row=1, column=1, value="B内容")
        xlsx = tmp_path / "sel.xlsx"
        _save_to_file(wb, xlsx)

        args = parse_args([str(xlsx), "-s", "SheetB"])
        code = run(args)
        assert code == 0
        md = xlsx.with_suffix(".md").read_text(encoding="utf-8")
        assert "B内容" in md
        assert "A内容" not in md

    def test_debug_mode_does_not_crash(self, tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
        xlsx = tmp_path / "debug.xlsx"
        _make_simple_xlsx(xlsx, value="デバッグ")
        args = parse_args([str(xlsx), "--debug"])
        code = run(args)
        assert code == 0
        captured = capsys.readouterr()
        # TextBlock の JSON が stderr に出力される
        assert "デバッグ" in captured.err

    def test_empty_sheet_is_skipped(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        ws.title = "Empty"
        # セルなし
        xlsx = tmp_path / "empty.xlsx"
        _save_to_file(wb, xlsx)

        args = parse_args([str(xlsx)])
        code = run(args)
        assert code == 0

    def test_all_empty_sheets_returns_zero(self, tmp_path: Path) -> None:
        wb = Workbook()
        xlsx = tmp_path / "allempty.xlsx"
        _save_to_file(wb, xlsx)
        args = parse_args([str(xlsx)])
        code = run(args)
        assert code == 0

    def test_heading_detected_with_base_font_size(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        cell = ws.cell(row=1, column=1, value="タイトル")
        cell.font = Font(size=18)
        xlsx = tmp_path / "font.xlsx"
        _save_to_file(wb, xlsx)

        args = parse_args([str(xlsx), "--base-font-size", "11.0"])
        code = run(args)
        assert code == 0
        md = xlsx.with_suffix(".md").read_text(encoding="utf-8")
        assert "# タイトル" in md


# ---------------------------------------------------------------------------
# run() 異常系
# ---------------------------------------------------------------------------


class TestRunErrors:
    def test_file_not_found_returns_1(self, tmp_path: Path) -> None:
        args = parse_args([str(tmp_path / "nonexistent.xlsx")])
        code = run(args)
        assert code == 1

    def test_unsupported_extension_returns_1(self, tmp_path: Path) -> None:
        f = tmp_path / "test.csv"
        f.write_text("a,b,c")
        args = parse_args([str(f)])
        code = run(args)
        assert code == 1

    def test_sheet_not_found_returns_1(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "doc.xlsx"
        _make_simple_xlsx(xlsx)
        args = parse_args([str(xlsx), "-s", "NoSuchSheet"])
        code = run(args)
        assert code == 1

    def test_sheet_index_out_of_range_returns_1(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "doc.xlsx"
        _make_simple_xlsx(xlsx)
        args = parse_args([str(xlsx), "-s", "99"])
        code = run(args)
        assert code == 1


# ---------------------------------------------------------------------------
# バッチ変換
# ---------------------------------------------------------------------------


class TestBatchConversion:
    def test_batch_converts_all_xlsx_in_dir(self, tmp_path: Path) -> None:
        """ディレクトリ指定で配下の全 xlsx を変換できること。"""
        _make_simple_xlsx(tmp_path / "a.xlsx", value="Aファイル")
        _make_simple_xlsx(tmp_path / "b.xlsx", value="Bファイル")
        args = parse_args([str(tmp_path)])
        code = run(args)
        assert code == 0
        assert (tmp_path / "a.md").exists()
        assert (tmp_path / "b.md").exists()
        assert "Aファイル" in (tmp_path / "a.md").read_text(encoding="utf-8")
        assert "Bファイル" in (tmp_path / "b.md").read_text(encoding="utf-8")

    def test_batch_empty_dir_returns_zero(self, tmp_path: Path) -> None:
        """変換対象がない場合も exit code 0 で終了すること。"""
        args = parse_args([str(tmp_path)])
        code = run(args)
        assert code == 0

    def test_batch_subdirectory_recursive(self, tmp_path: Path) -> None:
        """サブディレクトリ内の xlsx も変換されること。"""
        subdir = tmp_path / "sub"
        subdir.mkdir()
        _make_simple_xlsx(subdir / "c.xlsx", value="サブファイル")
        args = parse_args([str(tmp_path)])
        code = run(args)
        assert code == 0
        assert (subdir / "c.md").exists()


# ---------------------------------------------------------------------------
# ハイパーリンク変換
# ---------------------------------------------------------------------------


class TestHyperlinkConversion:
    def test_hyperlink_rendered_as_markdown_link(self, tmp_path: Path) -> None:
        """ハイパーリンクを持つセルが [text](url) 形式に変換されること。"""
        from openpyxl.styles import Font

        wb = Workbook()
        ws: Worksheet = wb.active  # type: ignore[assignment]
        cell = ws.cell(row=1, column=1, value="Anthropic")
        cell.hyperlink = "https://anthropic.com"
        cell.font = Font(color="0000FF", underline="single")
        xlsx = tmp_path / "link.xlsx"
        _save_to_file(wb, xlsx)

        args = parse_args([str(xlsx)])
        code = run(args)
        assert code == 0
        md = xlsx.with_suffix(".md").read_text(encoding="utf-8")
        assert "[Anthropic](https://anthropic.com)" in md

    def test_no_hyperlink_renders_plain_text(self, tmp_path: Path) -> None:
        """ハイパーリンクなしのセルは通常テキストとして出力されること。"""
        xlsx = tmp_path / "plain.xlsx"
        _make_simple_xlsx(xlsx, value="通常テキスト")
        args = parse_args([str(xlsx)])
        run(args)
        md = xlsx.with_suffix(".md").read_text(encoding="utf-8")
        assert "通常テキスト" in md
        assert "](http" not in md


# ---------------------------------------------------------------------------
# .xls 対応
# ---------------------------------------------------------------------------


class TestXlsConversion:
    xlwt = pytest.importorskip("xlwt")

    def _make_simple_xls(self, path: Path, value: str = "xlsテスト") -> None:
        wb = self.xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, value)
        wb.save(str(path))

    def _make_multi_sheet_xls(self, path: Path) -> None:
        wb = self.xlwt.Workbook()
        ws1 = wb.add_sheet("シート1")
        ws1.write(0, 0, "シート1の内容")
        ws2 = wb.add_sheet("シート2")
        ws2.write(0, 0, "シート2の内容")
        wb.save(str(path))

    def test_xls_basic_conversion(self, tmp_path: Path) -> None:
        """xls ファイルの基本変換が動作すること。"""
        xls = tmp_path / "test.xls"
        self._make_simple_xls(xls, value="xlsコンテンツ")
        args = parse_args([str(xls)])
        code = run(args)
        assert code == 0
        md = xls.with_suffix(".md").read_text(encoding="utf-8")
        assert "xlsコンテンツ" in md

    def test_xls_multi_sheet(self, tmp_path: Path) -> None:
        """xls の複数シートが統合されること。"""
        xls = tmp_path / "multi.xls"
        self._make_multi_sheet_xls(xls)
        args = parse_args([str(xls)])
        code = run(args)
        assert code == 0
        md = xls.with_suffix(".md").read_text(encoding="utf-8")
        assert "シート1の内容" in md
        assert "シート2の内容" in md

    def test_xls_sheet_by_name(self, tmp_path: Path) -> None:
        """xls のシート名指定が動作すること。"""
        xls = tmp_path / "sel.xls"
        self._make_multi_sheet_xls(xls)
        args = parse_args([str(xls), "-s", "シート2"])
        code = run(args)
        assert code == 0
        md = xls.with_suffix(".md").read_text(encoding="utf-8")
        assert "シート2の内容" in md
        assert "シート1の内容" not in md

    def test_xls_sheet_by_index(self, tmp_path: Path) -> None:
        """xls のシートインデックス指定が動作すること。"""
        xls = tmp_path / "idx.xls"
        self._make_multi_sheet_xls(xls)
        args = parse_args([str(xls), "-s", "0"])
        code = run(args)
        assert code == 0
        md = xls.with_suffix(".md").read_text(encoding="utf-8")
        assert "シート1の内容" in md

    def test_xls_sheet_not_found(self, tmp_path: Path) -> None:
        """存在しないシート名を指定した場合 exit code 1 を返すこと。"""
        xls = tmp_path / "err.xls"
        self._make_simple_xls(xls)
        args = parse_args([str(xls), "-s", "存在しないシート"])
        code = run(args)
        assert code == 1

    def test_xls_sheet_index_out_of_range(self, tmp_path: Path) -> None:
        xls = tmp_path / "err.xls"
        self._make_simple_xls(xls)
        args = parse_args([str(xls), "-s", "99"])
        code = run(args)
        assert code == 1

    def test_xls_bold_font(self, tmp_path: Path) -> None:
        """xls の太字セルが見出しとして変換されること。"""
        wb = self.xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        bold_style = self.xlwt.easyxf("font: bold true")
        ws.write(0, 0, "太字の見出し", bold_style)
        xls = tmp_path / "bold.xls"
        wb.save(str(xls))
        args = parse_args([str(xls)])
        code = run(args)
        assert code == 0
        md = xls.with_suffix(".md").read_text(encoding="utf-8")
        assert "太字の見出し" in md

    def test_batch_includes_xls(self, tmp_path: Path) -> None:
        """バッチ変換で xls も対象になること。"""
        self._make_simple_xls(tmp_path / "data.xls", value="xlsバッチ")
        args = parse_args([str(tmp_path)])
        code = run(args)
        assert code == 0
        assert (tmp_path / "data.md").exists()


# ---------------------------------------------------------------------------
# _convert_file の追加エラーケース
# ---------------------------------------------------------------------------


class TestConvertFileErrors:
    def test_permission_error_in_batch(self, tmp_path: Path) -> None:
        """バッチ変換中のファイルでエラーが起きても他のファイルの変換は継続すること。"""
        _make_simple_xlsx(tmp_path / "ok.xlsx", value="正常ファイル")
        # 読み込めないファイル（壊れた xlsx）を追加
        broken = tmp_path / "broken.xlsx"
        broken.write_bytes(b"not a valid xlsx")
        args = parse_args([str(tmp_path)])
        # エラーは起きるかもしれないが ok.md は生成される
        run(args)
        assert (tmp_path / "ok.md").exists()
