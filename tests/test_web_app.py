"""Web API (FastAPI) のテスト。

httpx の AsyncClient ではなく TestClient（同期）を使用する。
"""

from __future__ import annotations

import io
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from excel_to_markdown.cli import run_file
from excel_to_markdown.web.app import create_app


# ---------------------------------------------------------------------------
# フィクスチャ
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(text: str = "テスト") -> bytes:
    """最小限の .xlsx ファイルをメモリ上で生成してバイト列を返す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=text)  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def client() -> TestClient:
    app = create_app()
    return TestClient(app)


@pytest.fixture(scope="module")
def xlsx_bytes() -> bytes:
    return _make_xlsx_bytes("Hello World")


@pytest.fixture
def tmp_xlsx(tmp_path: Path) -> Path:
    """一時ディレクトリに最小 .xlsx を書き出して返す。"""
    p = tmp_path / "test.xlsx"
    p.write_bytes(_make_xlsx_bytes("run_file テスト"))
    return p


# ---------------------------------------------------------------------------
# GET /health
# ---------------------------------------------------------------------------


class TestHealth:
    def test_health_ok(self, client: TestClient) -> None:
        res = client.get("/health")
        assert res.status_code == 200
        assert res.json() == {"status": "ok"}


# ---------------------------------------------------------------------------
# GET /
# ---------------------------------------------------------------------------


class TestIndex:
    def test_index_returns_html(self, client: TestClient) -> None:
        res = client.get("/")
        assert res.status_code == 200
        assert "text/html" in res.headers["content-type"]
        assert "D" in res.text  # D&D UI が含まれる


# ---------------------------------------------------------------------------
# POST /api/convert — 単一ファイル
# ---------------------------------------------------------------------------


class TestConvertSingle:
    def test_single_xlsx_returns_markdown(
        self, client: TestClient, xlsx_bytes: bytes
    ) -> None:
        res = client.post(
            "/api/convert",
            files=[("files", ("test.xlsx", xlsx_bytes, "application/octet-stream"))],
        )
        assert res.status_code == 200
        assert "text/markdown" in res.headers["content-type"]
        assert "Hello World" in res.text

    def test_content_disposition_has_md_filename(
        self, client: TestClient, xlsx_bytes: bytes
    ) -> None:
        res = client.post(
            "/api/convert",
            files=[("files", ("myfile.xlsx", xlsx_bytes, "application/octet-stream"))],
        )
        assert res.status_code == 200
        cd = res.headers.get("content-disposition", "")
        assert "myfile.md" in cd

    def test_unsupported_extension_returns_400(self, client: TestClient) -> None:
        res = client.post(
            "/api/convert",
            files=[("files", ("data.csv", b"a,b\n1,2", "text/csv"))],
        )
        assert res.status_code == 400
        assert "対応していないファイル形式" in res.json()["detail"]


# ---------------------------------------------------------------------------
# POST /api/convert — 複数ファイル
# ---------------------------------------------------------------------------


class TestConvertMultiple:
    def test_multiple_xlsx_returns_zip(self, client: TestClient) -> None:
        file1 = _make_xlsx_bytes("ファイル1")
        file2 = _make_xlsx_bytes("ファイル2")
        res = client.post(
            "/api/convert",
            files=[
                ("files", ("a.xlsx", file1, "application/octet-stream")),
                ("files", ("b.xlsx", file2, "application/octet-stream")),
            ],
        )
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/zip"

        zf = zipfile.ZipFile(io.BytesIO(res.content))
        names = zf.namelist()
        assert "a.md" in names
        assert "b.md" in names

    def test_partial_error_returns_zip_with_header(self, client: TestClient) -> None:
        """一部ファイルが非対応拡張子でも成功分を ZIP で返す。"""
        xlsx = _make_xlsx_bytes("成功ファイル")
        res = client.post(
            "/api/convert",
            files=[
                ("files", ("good.xlsx", xlsx, "application/octet-stream")),
                ("files", ("bad.csv", b"a,b", "text/csv")),
            ],
        )
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/zip"
        assert "bad.csv" in res.headers.get("x-conversion-errors", "")

        zf = zipfile.ZipFile(io.BytesIO(res.content))
        assert "good.md" in zf.namelist()

    def test_all_fail_returns_422(self, client: TestClient) -> None:
        res = client.post(
            "/api/convert",
            files=[
                ("files", ("bad1.csv", b"a,b", "text/csv")),
                ("files", ("bad2.txt", b"hello", "text/plain")),
            ],
        )
        assert res.status_code == 422


# ---------------------------------------------------------------------------
# run_file() ユニットテスト
# ---------------------------------------------------------------------------


class TestRunFile:
    def test_run_file_returns_markdown(self, tmp_xlsx: Path) -> None:
        md = run_file(tmp_xlsx)
        assert isinstance(md, str)
        assert "run_file テスト" in md

    def test_run_file_invalid_extension_raises(self, tmp_path: Path) -> None:
        p = tmp_path / "dummy.csv"
        p.write_text("a,b")
        with pytest.raises(ValueError, match="対応していないファイル形式"):
            run_file(p)

    def test_run_file_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            run_file(tmp_path / "not_exist.xlsx")

    def test_run_file_empty_sheet_returns_empty_string(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        p = tmp_path / "empty.xlsx"
        wb.save(str(p))
        result = run_file(p)
        assert result == ""

    def test_run_file_multiple_sheets(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"  # type: ignore[union-attr]
        ws1.cell(row=1, column=1, value="シート1")  # type: ignore[union-attr]
        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(row=1, column=1, value="シート2")
        p = tmp_path / "multi.xlsx"
        wb.save(str(p))

        md = run_file(p)
        assert "Sheet1" in md
        assert "Sheet2" in md
        assert "シート1" in md
        assert "シート2" in md
