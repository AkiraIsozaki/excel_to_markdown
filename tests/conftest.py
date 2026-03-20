"""テスト共通フィクスチャ。

openpyxl で xlsx ワークブックをプログラム的に生成するヘルパーを提供する。
バイナリの .xlsx ファイルはリポジトリに含めず、テスト実行時にオンメモリ生成する。
"""

from __future__ import annotations

import io
from typing import Callable

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@pytest.fixture
def make_workbook() -> Callable[[], Workbook]:
    """空の Workbook を生成するファクトリを返す。"""

    def _factory() -> Workbook:
        return Workbook()

    return _factory


@pytest.fixture
def save_and_reload() -> Callable[[Workbook], Workbook]:
    """Workbook をオンメモリで保存・再読み込みして返すファクトリを返す。

    openpyxl でセルを書き込んだ後、data_only=True で再読み込みすることで
    計算済み値を取得する際に使用する。
    """
    import openpyxl

    def _reload(wb: Workbook) -> Workbook:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)

    return _reload


def make_simple_sheet(wb: Workbook, data: list[tuple[int, int, str | None]]) -> Worksheet:
    """指定セルにテキストを書き込んだシートを作成する。

    data: [(row, col, value), ...]
    """
    ws: Worksheet = wb.active  # type: ignore[assignment]
    for r, c, v in data:
        ws.cell(row=r, column=c, value=v)
    return ws
