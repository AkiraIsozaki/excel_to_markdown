"""xlsx_reader.py のユニットテスト。"""

from __future__ import annotations

import io

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_to_markdown.reader.xlsx_reader import (
    extract_bg_color,
    extract_font_props,
    get_print_area,
    read_sheet,
)


def _reload(wb: Workbook) -> Workbook:
    """Workbook をオンメモリ保存・再読み込みする。"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


# ---------------------------------------------------------------------------
# read_sheet: 基本動作
# ---------------------------------------------------------------------------


class TestReadSheetBasic:
    def test_single_cell(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Hello")  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        assert len(cells) == 1
        assert cells[0].value == "Hello"
        assert cells[0].row == 1
        assert cells[0].col == 1

    def test_empty_cells_skipped(self) -> None:
        """値なしセルは RawCell に含まれるが value=None となる。"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="A")  # type: ignore[union-attr]
        ws.cell(row=1, column=2, value=None)  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        # 値ありセルだけ RawCell が生成されることを確認
        values = [c.value for c in cells if c.value is not None]
        assert values == ["A"]

    def test_numeric_converted_to_str(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=42)  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        non_none = [c for c in cells if c.value is not None]
        assert non_none[0].value == "42"


# ---------------------------------------------------------------------------
# read_sheet: 結合セル
# ---------------------------------------------------------------------------


class TestReadSheetMergedCells:
    def test_merge_origin_has_value(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Merged")  # type: ignore[union-attr]
        ws.merge_cells("A1:B2")  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        origin = next(c for c in cells if c.row == 1 and c.col == 1)
        assert origin.is_merge_origin is True
        assert origin.merge_row_span == 2
        assert origin.merge_col_span == 2
        assert origin.value == "Merged"

    def test_merge_non_origin_value_is_none(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Merged")  # type: ignore[union-attr]
        ws.merge_cells("A1:B2")  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        non_origins = [c for c in cells if not c.is_merge_origin and (c.row, c.col) != (1, 1)]
        for c in non_origins:
            assert c.value is None
            assert c.is_merge_origin is False


# ---------------------------------------------------------------------------
# read_sheet: 印刷領域
# ---------------------------------------------------------------------------


class TestReadSheetPrintArea:
    def test_cells_outside_print_area_excluded(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Inside")  # type: ignore[union-attr]
        ws.cell(row=5, column=5, value="Outside")  # type: ignore[union-attr]
        ws.print_area = "A1:C3"  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        values = [c.value for c in cells if c.value is not None]
        assert "Inside" in values
        assert "Outside" not in values


# ---------------------------------------------------------------------------
# read_sheet: 非表示行・列
# ---------------------------------------------------------------------------


class TestReadSheetHidden:
    def test_hidden_row_excluded(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Visible")  # type: ignore[union-attr]
        ws.cell(row=2, column=1, value="Hidden")  # type: ignore[union-attr]
        ws.row_dimensions[2].hidden = True  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        values = [c.value for c in cells if c.value is not None]
        assert "Visible" in values
        assert "Hidden" not in values

    def test_hidden_col_excluded(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Visible")  # type: ignore[union-attr]
        ws.cell(row=1, column=2, value="Hidden")  # type: ignore[union-attr]
        ws.column_dimensions[get_column_letter(2)].hidden = True  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cells = read_sheet(wb2.active)  # type: ignore[arg-type]
        values = [c.value for c in cells if c.value is not None]
        assert "Visible" in values
        assert "Hidden" not in values


# ---------------------------------------------------------------------------
# extract_font_props
# ---------------------------------------------------------------------------


class TestExtractFontProps:
    def test_bold_italic(self) -> None:
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="X")  # type: ignore[union-attr]
        cell.font = Font(bold=True, italic=True, size=14)
        wb2 = _reload(wb)
        cell2 = wb2.active.cell(row=1, column=1)  # type: ignore[union-attr]
        bold, italic, strike, underline, size, color = extract_font_props(cell2)
        assert bold is True
        assert italic is True
        assert strike is False
        assert size == 14.0

    def test_no_font(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="X")  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cell2 = wb2.active.cell(row=1, column=1)  # type: ignore[union-attr]
        bold, italic, strike, underline, size, color = extract_font_props(cell2)
        assert bold is False
        assert size is None or isinstance(size, float)


# ---------------------------------------------------------------------------
# extract_bg_color
# ---------------------------------------------------------------------------


class TestExtractBgColor:
    def test_solid_fill(self) -> None:
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="X")  # type: ignore[union-attr]
        cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        wb2 = _reload(wb)
        cell2 = wb2.active.cell(row=1, column=1)  # type: ignore[union-attr]
        color = extract_bg_color(cell2)
        assert color == "FFFF0000"

    def test_no_fill_returns_none(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="X")  # type: ignore[union-attr]
        wb2 = _reload(wb)
        cell2 = wb2.active.cell(row=1, column=1)  # type: ignore[union-attr]
        assert extract_bg_color(cell2) is None


# ---------------------------------------------------------------------------
# get_print_area
# ---------------------------------------------------------------------------


class TestGetPrintArea:
    def test_no_print_area(self) -> None:
        wb = Workbook()
        wb2 = _reload(wb)
        assert get_print_area(wb2.active) is None  # type: ignore[arg-type]

    def test_with_print_area(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.print_area = "B2:E10"  # type: ignore[union-attr]
        wb2 = _reload(wb)
        result = get_print_area(wb2.active)  # type: ignore[arg-type]
        assert result == (2, 2, 10, 5)
