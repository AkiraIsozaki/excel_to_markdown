"""
実際の Excel 方眼紙ドキュメントを生成するスクリプト。

SIer の要件定義書・基本設計書によくある構成を再現:
  - 方眼紙設定（全列幅 2.5 文字幅に統一）
  - タイトル（大フォント・セル結合）
  - プロジェクト情報（ラベル:値パターン）
  - セクション見出し（着色ヘッダー・14pt Bold）
  - 機能一覧テーブル（5列）
  - 詳細要件リスト（インデント付き）
  - 非機能要件テーブル
  - セルコメント（脚注変換の確認用）
  - 複数シート（機能要件 / 画面一覧）

使い方:
    python tests/e2e/fixtures/make_sample_houganshi.py
    # → tests/e2e/fixtures/sample_houganshi.xlsx を生成
"""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "sample_houganshi.xlsx"


def build_workbook() -> Workbook:
    wb = Workbook()
    _build_sheet1(wb)
    _build_sheet2(wb)
    return wb


def _build_sheet1(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "機能要件"

    # 方眼紙設定: 全列幅を 2.5 文字幅に統一
    for i in range(1, 61):
        ws.column_dimensions[get_column_letter(i)].width = 2.5
    for i in range(1, 100):
        ws.row_dimensions[i].height = 15

    blue_fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
    gray_fill = PatternFill(fill_type="solid", fgColor="FFCCCCCC")
    header_fill = PatternFill(fill_type="solid", fgColor="FFD9E1F2")

    # --- タイトル (row2) ---
    ws.merge_cells("B2:AJ2")
    c = ws["B2"]
    c.value = "機能要件定義書"
    c.font = Font(name="MS Gothic", size=18, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    # --- プロジェクト情報 (row4-7) ---
    info = [
        (4, "プロジェクト名", "在庫管理システム刷新プロジェクト"),
        (5, "作成者",         "田中 誠一"),
        (6, "作成日",         "2025-01-15"),
        (7, "バージョン",     "1.0"),
    ]
    for row, label, value in info:
        ws.merge_cells(f"B{row}:H{row}")
        ws.merge_cells(f"I{row}:R{row}")
        lc = ws[f"B{row}"]
        lc.value = label
        lc.font = Font(name="MS Gothic", bold=True)
        lc.fill = gray_fill
        vc = ws[f"I{row}"]
        vc.value = value
        vc.font = Font(name="MS Gothic")

    # --- セクション1: 概要 (row9) ---
    _section_header(ws, 9, "1. 概要", blue_fill)
    _paragraph(ws, 10, "B", "本書は在庫管理システム刷新プロジェクトにおける機能要件を定義するものである。")
    _paragraph(ws, 11, "B", "本システムは既存の在庫管理業務を効率化し、リアルタイムな在庫状況の把握を可能にする。")

    # --- セクション2: 機能一覧 (row13) ---
    _section_header(ws, 13, "2. 機能一覧", blue_fill)
    t_headers = ["No.", "機能名", "概要", "優先度", "備考"]
    t_cs = [2, 5, 12, 30, 35]
    t_ce = [4, 11, 29, 34, 42]
    _table_header_row(ws, 14, t_headers, t_cs, t_ce, header_fill)
    table_data = [
        ("1", "在庫照会",     "商品コード・カテゴリ・倉庫別に在庫数量を照会できる",             "高", ""),
        ("2", "入庫登録",     "仕入先からの入荷情報を登録し、在庫数を更新する",                 "高", "バーコード対応"),
        ("3", "出庫登録",     "出荷指示に基づき出庫処理を行い、在庫数を減算する",               "高", ""),
        ("4", "在庫調整",     "棚卸結果に基づき在庫数量を修正する",                             "中", "承認フロー必要"),
        ("5", "アラート通知", "在庫数が設定閾値を下回った場合にメール通知を送信する",           "中", "要件#12参照"),
    ]
    for i, row_data in enumerate(table_data):
        _table_data_row(ws, 15 + i, row_data, t_cs, t_ce)

    # --- セクション3: 詳細要件 (row21) ---
    _section_header(ws, 21, "3. 詳細要件", blue_fill)
    _subsection_header(ws, 23, "3.1 在庫照会")
    reqs1 = [
        (25, 4, "商品コード（JANコード）で在庫を検索できること"),
        (26, 4, "カテゴリー、倉庫、ロケーションで絞り込み検索できること"),
        (27, 4, "検索結果は最大1000件まで表示し、ページング機能を提供すること"),
        (28, 4, "CSV形式でエクスポートできること"),
        (29, 6, "エクスポート対象は検索結果全件または選択した行のみを選択できること"),
        (30, 6, "エクスポートファイルのフォーマットはシステム設定で変更可能とする"),
    ]
    for row, col, text in reqs1:
        _paragraph(ws, row, get_column_letter(col), text)

    _subsection_header(ws, 32, "3.2 入庫登録")
    reqs2 = [
        (34, 4, "入庫伝票番号、入庫日、仕入先、商品情報を入力できること"),
        (35, 4, "バーコードリーダーによる商品コードのスキャン入力に対応すること"),
        (36, 4, "入庫確定後は在庫マスタに自動反映されること"),
        (37, 4, "入庫履歴は最低5年間保持すること"),
    ]
    for row, col, text in reqs2:
        _paragraph(ws, row, get_column_letter(col), text)

    # --- セクション4: 非機能要件 (row39) ---
    _section_header(ws, 39, "4. 非機能要件", blue_fill)
    nfr_headers = ["分類", "要件", "目標値"]
    nfr_cs = [2, 8, 28]
    nfr_ce = [7, 27, 42]
    _table_header_row(ws, 41, nfr_headers, nfr_cs, nfr_ce, header_fill)
    nfr_data = [
        ("性能",         "画面の初期表示時間",   "3秒以内"),
        ("性能",         "在庫検索のレスポンス", "1秒以内（1000件以下）"),
        ("可用性",       "稼働率",               "99.5%以上（計画停止を除く）"),
        ("セキュリティ", "認証方式",             "ID/パスワード + 2要素認証"),
        ("保守性",       "ログ保持期間",         "操作ログ: 1年間、エラーログ: 3年間"),
    ]
    for i, row_data in enumerate(nfr_data):
        _table_data_row(ws, 42 + i, row_data, nfr_cs, nfr_ce)

    # セルコメント付き備考
    ws.merge_cells("B47:AJ47")
    c47 = ws["B47"]
    c47.value = "※ 稼働率の計測は毎月末に実施し、SLAレポートとして提出する"
    c47.font = Font(name="MS Gothic", italic=True)
    c47.comment = Comment("SLAレポートのテンプレートは別途資料参照。\n担当: 佐藤PM", "田中")


def _build_sheet2(wb: Workbook) -> None:
    ws = wb.create_sheet("画面一覧")
    for i in range(1, 61):
        ws.column_dimensions[get_column_letter(i)].width = 2.5
    for i in range(1, 50):
        ws.row_dimensions[i].height = 15

    header_fill = PatternFill(fill_type="solid", fgColor="FFD9E1F2")

    ws.merge_cells("B2:AJ2")
    t = ws["B2"]
    t.value = "画面一覧"
    t.font = Font(name="MS Gothic", size=16, bold=True)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 25

    s_headers = ["画面ID", "画面名", "概要", "遷移元"]
    s_cs = [2, 6, 12, 32]
    s_ce = [5, 11, 31, 42]
    _table_header_row(ws, 4, s_headers, s_cs, s_ce, header_fill)

    screens = [
        ("SCR-001", "ログイン画面",   "ユーザーIDとパスワードでシステムにログインする", "（初期画面）"),
        ("SCR-010", "メニュー画面",   "各機能へのナビゲーションを提供するトップ画面",   "SCR-001"),
        ("SCR-101", "在庫照会画面",   "条件を指定して在庫情報を検索・表示する",         "SCR-010"),
        ("SCR-201", "入庫登録画面",   "入庫情報を入力し在庫数を更新する",               "SCR-010"),
        ("SCR-301", "出庫登録画面",   "出庫情報を入力し在庫数を減算する",               "SCR-010"),
    ]
    for i, row_data in enumerate(screens):
        _table_data_row(ws, 5 + i, row_data, s_cs, s_ce)


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def _section_header(ws, row: int, text: str, fill: PatternFill) -> None:
    ws.merge_cells(f"B{row}:AJ{row}")
    c = ws[f"B{row}"]
    c.value = text
    c.font = Font(name="MS Gothic", size=14, bold=True, color="FFFFFFFF")
    c.fill = fill
    ws.row_dimensions[row].height = 20


def _subsection_header(ws, row: int, text: str) -> None:
    ws.merge_cells(f"B{row}:AJ{row}")
    c = ws[f"B{row}"]
    c.value = text
    c.font = Font(name="MS Gothic", size=12, bold=True)


def _paragraph(ws, row: int, start_col: str, text: str) -> None:
    ws.merge_cells(f"{start_col}{row}:AJ{row}")
    c = ws[f"{start_col}{row}"]
    c.value = text
    c.font = Font(name="MS Gothic")


def _table_header_row(
    ws, row: int, headers: list[str],
    col_starts: list[int], col_ends: list[int],
    fill: PatternFill,
) -> None:
    for header, cs, ce in zip(headers, col_starts, col_ends):
        ws.merge_cells(f"{get_column_letter(cs)}{row}:{get_column_letter(ce)}{row}")
        c = ws[f"{get_column_letter(cs)}{row}"]
        c.value = header
        c.font = Font(name="MS Gothic", bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")


def _table_data_row(
    ws, row: int, values: tuple[str, ...],
    col_starts: list[int], col_ends: list[int],
) -> None:
    for val, cs, ce in zip(values, col_starts, col_ends):
        ws.merge_cells(f"{get_column_letter(cs)}{row}:{get_column_letter(ce)}{row}")
        c = ws[f"{get_column_letter(cs)}{row}"]
        c.value = val
        c.font = Font(name="MS Gothic")


if __name__ == "__main__":
    wb = build_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    OUTPUT_PATH.write_bytes(buf.getvalue())
    print(f"Generated: {OUTPUT_PATH}")
