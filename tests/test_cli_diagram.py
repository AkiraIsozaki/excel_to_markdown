"""CLI --diagram フラグ、および自動統合変換のテスト。

test_cli.py は xlwt の importorskip でファイル全体がスキップされるケースがあるため、
diagram 関連の CLI テストはこのファイルに分離する。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from excel_to_markdown.cli import parse_args, run

SAMPLE_FLOWCHART = Path(__file__).parent / "e2e/fixtures/sample_flowchart.xlsx"
SAMPLE_1 = Path(__file__).parent / "e2e/fixtures/1.xlsx"
SAMPLE_GYOUMU = Path(__file__).parent / "e2e/fixtures/gyoumuflow_answer.xlsx"


class TestDiagramFlag:
    def test_diagram_flag_parses(self) -> None:
        args = parse_args(["dummy.xlsx", "--diagram"])
        assert args.diagram is True

    def test_diagram_default_false(self) -> None:
        args = parse_args(["dummy.xlsx"])
        assert args.diagram is False

    def test_diagram_run_outputs_mermaid(
        self, capsys: pytest.CaptureFixture[str]
    ) -> None:
        if not SAMPLE_FLOWCHART.exists():
            pytest.skip("サンプルファイルが見つかりません")
        args = parse_args([str(SAMPLE_FLOWCHART), "--diagram"])
        code = run(args)
        assert code == 0
        captured = capsys.readouterr()
        assert "```mermaid" in captured.out
        assert "flowchart" in captured.out
        assert "開始" in captured.out

    def test_diagram_run_saves_to_file(self, tmp_path: Path) -> None:
        if not SAMPLE_FLOWCHART.exists():
            pytest.skip("サンプルファイルが見つかりません")
        out_path = tmp_path / "diagram.md"
        args = parse_args(
            [str(SAMPLE_FLOWCHART), "--diagram", "--output", str(out_path)]
        )
        code = run(args)
        assert code == 0
        assert out_path.exists()
        content = out_path.read_text(encoding="utf-8")
        assert "```mermaid" in content

    def test_diagram_no_drawing_returns_zero(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "no_drawing.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx)
        args = parse_args([str(xlsx), "--diagram"])
        code = run(args)
        assert code == 0


# ---------------------------------------------------------------------------
# 自動統合変換テスト（--diagram なし、通常変換での自動Mermaid混在）
# ---------------------------------------------------------------------------


class TestAutoMermaidIntegration:
    """drawingを持つxlsxを通常変換すると自動でMermaidブロックが挿入されることを検証。"""

    def test_sample_flowchart_auto_mermaid(self, tmp_path: Path) -> None:
        """sample_flowchart.xlsx は drawing → 通常変換でもMermaid出力になる。"""
        if not SAMPLE_FLOWCHART.exists():
            pytest.skip("サンプルファイルが見つかりません")
        out = tmp_path / "out.md"
        args = parse_args([str(SAMPLE_FLOWCHART), "--output", str(out)])
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" in content
        assert "flowchart" in content

    def test_1xlsx_drawing_sheet_has_mermaid(self, tmp_path: Path) -> None:
        """1.xlsx の '画面遷移' シートはMermaidを含む。"""
        if not SAMPLE_1.exists():
            pytest.skip("1.xlsx が見つかりません")
        out = tmp_path / "out.md"
        args = parse_args([str(SAMPLE_1), "--output", str(out)])
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" in content
        # ログインフローの主要ノードが含まれる
        assert "ログイン画面" in content
        assert "トップ画面" in content

    def test_1xlsx_no_drawing_sheet_normal_markdown(self, tmp_path: Path) -> None:
        """1.xlsx の '共通要件' シートは通常markdownのみ（Mermaidなし）。"""
        if not SAMPLE_1.exists():
            pytest.skip("1.xlsx が見つかりません")
        # シートを指定して変換
        out = tmp_path / "out.md"
        args = parse_args([str(SAMPLE_1), "--sheet", "共通要件", "--output", str(out)])
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" not in content
        assert "共通要件" in content

    def test_gyoumu_複数部署_has_mermaid(self, tmp_path: Path) -> None:
        """gyoumuflow_answer.xlsx の '複数部署' シートはMermaidを含む。"""
        if not SAMPLE_GYOUMU.exists():
            pytest.skip("gyoumuflow_answer.xlsx が見つかりません")
        out = tmp_path / "out.md"
        args = parse_args(
            [str(SAMPLE_GYOUMU), "--sheet", "複数部署", "--output", str(out)]
        )
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" in content
        assert "受付用紙に記入する" in content

    def test_gyoumu_複数部署_has_swimlane_subgraphs(self, tmp_path: Path) -> None:
        """gyoumuflow_answer.xlsx の '複数部署' シートはスイムレーン subgraph を含む。"""
        if not SAMPLE_GYOUMU.exists():
            pytest.skip("gyoumuflow_answer.xlsx が見つかりません")
        out = tmp_path / "out.md"
        args = parse_args(
            [str(SAMPLE_GYOUMU), "--sheet", "複数部署", "--output", str(out)]
        )
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "subgraph" in content
        assert "お客様" in content
        assert "〇〇ガラス店" in content
        assert "問屋" in content

    def test_gyoumu_はじめに_no_mermaid(self, tmp_path: Path) -> None:
        """gyoumuflow_answer.xlsx の 'はじめに' シートはMermaidなし。"""
        if not SAMPLE_GYOUMU.exists():
            pytest.skip("gyoumuflow_answer.xlsx が見つかりません")
        out = tmp_path / "out.md"
        args = parse_args(
            [str(SAMPLE_GYOUMU), "--sheet", "はじめに", "--output", str(out)]
        )
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" not in content
        assert "業務フロー" in content

    def test_no_drawing_xlsx_no_mermaid(self, tmp_path: Path) -> None:
        """drawingなしのxlsxは通常markdownのみ。"""
        xlsx = tmp_path / "plain.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "テキスト"  # type: ignore[index]
        wb.save(xlsx)
        out = tmp_path / "out.md"
        args = parse_args([str(xlsx), "--output", str(out)])
        code = run(args)
        assert code == 0
        content = out.read_text(encoding="utf-8")
        assert "```mermaid" not in content
        assert "テキスト" in content
